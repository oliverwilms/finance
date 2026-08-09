# excel_postgres.py
#  
# 
import openpyxl
import os
import psycopg2
import sys

def get_connection():
    return psycopg2.connect(
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        host=os.getenv("DB_HOST"),
        port=os.getenv("DB_PORT")
    )

def iterate_excel_all_sheets(file_path):
    """
    Iterates over all sheets in an Excel file and prints non-empty cells.
    :param file_path: Path to the Excel file (.xlsx)
    """
    try:
        # Validate file existence
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # Validate file extension
        _, ext = os.path.splitext(file_path)
        if not ext in ('.xlsx', '.xlsm'):
            raise ValueError("Only .xlsx and xlsm files are supported.")

        # Initialize postgres connection
        conn = get_connection()
        cur = conn.cursor()

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Loop through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")

            # Iterate over all rows and cells
            for row in sheet.iter_rows():
                cur.execute(
                    "INSERT INTO transactions (date, checknum, description, vendor, category, subcategory, account, credit, debit, status) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (row["Date"], row["Check"], row["Memo"], row["Merchant"], row["Category"], row["SubCategory"], row["Account"], row["Credit"], row["Debit"], row["Status"])
                )

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    # Example usage
    excel_file = r"C:\Users\olive\OneDrive\12NEW\Miles_2026.xlsx"
    iterate_excel_all_sheets(excel_file)
